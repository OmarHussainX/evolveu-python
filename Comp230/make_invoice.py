import openpyxl


def make_invoice(invoice_id):
    """
    Receives an integer (invoice id) as parameter, obtains
    invoice data from a spreadhseet ('sales_data.xlsx') and
    produces an invoice with the following format:

    Andrew Wang (client ID: 124)
                                           Invoice #  266       
                                        Invoice date  2019-04-30


      Quantity  Item                 Unit price       Line total
      --------  ----                 ----------       ----------
         9      Pen                       70.00       $   630.00
         2      Stapler                   40.00       $    80.00
         1      Binder                    10.00       $    10.00


                                                      ----------
                                          TOTAL       $   720.00
    """
    # - Open 'sales_data.xlsx', obtain iterators for worksheets of interest
    # - Use 'invoice_id' to obtain customer id and invoice date
    # - Use customer id to obtain customer's name
    #
    # Create invoice:
    # - add header showing customer name, id and invoice id, date
    # - initialise invoice total to 0
    # - loop through line items for the invoice with id 'invoice_id'
    #       * determine product id, units sold from the line item
    #       * use product id to obtain product name, unit price
    #       * calculate line total (units sold * unit price)
    #       * add line total to invoice total
    #       * add a line to invoice showing units sold, product name,
    #         unit price, and line total
    # - add footer showing invoice total

    wb = openpyxl.load_workbook('sales_data.xlsx')

    customers_ws = wb['customers']
    invoices_ws = wb['invoices']
    line_items_ws = wb['invoice line items']
    products_ws = wb['products']

    invoice_txt = ''        # stores text of the entire invoice
    customer_name = None
    customer_id = None
    invoice_date = None
    invoice_total = 0
    products = {}           # dictionary of products
    line_items = {}         # dictionary of invoice line items

    # Obtain customer id and invoice date from 'invoices' worksheet
    for i in range(2, invoices_ws.max_row + 1):
            if not (invoices_ws.cell(i, 1).value is None):
                if (invoice_id == invoices_ws.cell(i, 1).value):
                    customer_id = invoices_ws.cell(i, 3).value
                    invoice_date = invoices_ws.cell(i, 2).value\
                        .strftime('%Y-%m-%d')

    if customer_id is None:
        return

    # Obtain customer name from 'customers' worksheet
    for i in range(2, customers_ws.max_row + 1):
            if not (customers_ws.cell(i, 1).value is None):
                if (customer_id == customers_ws.cell(i, 1).value):
                    customer_name = customers_ws.cell(i, 2).value + \
                        ' ' + customers_ws.cell(i, 3).value

    # Iterate over rows in 'products' worksheet, build dictionary of
    # product name and unit price, with product ID as key
    for i in range(2, products_ws.max_row + 1):
            if not (products_ws.cell(i, 1).value is None):
                products[products_ws.cell(i, 1).value] = \
                    {'name': products_ws.cell(i, 2).value,
                     'price': products_ws.cell(i, 3).value}

    # Iterate over rows in 'invoice line items' worksheet, build dictionary of
    # units sold, with product ID as key
    for i in range(2, line_items_ws.max_row + 1):
        if not (line_items_ws.cell(i, 1).value is None):
            if (invoice_id == line_items_ws.cell(i, 2).value):
                line_items[line_items_ws.cell(i, 3).value] = \
                    line_items_ws.cell(i, 4).value

    invoice_txt += f'\
{customer_name} (client ID: {customer_id})\n\
{format("Invoice #  ", ">50")}\
{format(invoice_id, "<10")}\n\
{format("Invoice date  ", ">50")}\
{invoice_date}\n'

    """
    Invoice is 60 columns wide:
    12 cols: Quantity, centered
    20 cols: Item name, left-aligned
    11 cols: Unit price, right-aligned
    17 cols: Line total, right-aligned
    """
    invoice_txt += f'\n\n\
{format("Quantity", "^12")}\
{format("Item", "<20")}\
{format("Unit price", ">11")}\
{format("Line total", ">17")}\n\
{format("--------", "^12")}\
{format("----", "<20")}\
{format("----------", ">11")}\
{format("----------", ">17")}\n'

    for prod_id, units_sold in line_items.items():
        line_total = units_sold * products[prod_id]['price']
        invoice_total += line_total
        invoice_txt += f'\
{format(units_sold, "^12")}\
{format(products[prod_id]["name"], "20")}\
{format(products[prod_id]["price"], ">11,.2f")}\
{format("$", ">8")}\
{format(line_total, ">9,.2f")}\n'

    invoice_txt += f'\n\n\
{format(" ", ">49")}\
{format("----------", ">11")}\n\
{format("TOTAL", ">43")}\
{format("$", ">8")}\
{format(invoice_total, "9,.2f")}\n'

    print(invoice_txt)


def main():
    # print(f'\n---------- {__file__} ----------')
    make_invoice(266)


if __name__ == '__main__':
    main()
