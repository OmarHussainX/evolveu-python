"""
TODO

1: check for empty rows while iterating, and skip
   (this will allows for the possibility of empty rows
   interspersed amongst populated ones)

2: unique ID for line items sheet

3: re-arrange column order: primary key always first
"""

import openpyxl

wk = openpyxl.load_workbook('sales_data.xlsx')

inv_wksheet = wk['Invoices']
line_items_wksheet = wk['Inv Line Items']
products_wksheet = wk['Product']


# Issue with using max_row as iteration limit:
#    spreadsheet has *several* empty rows...
rows = inv_wksheet.max_row
columns = inv_wksheet.max_column


# Iterate over rows in worksheet, build dictionary of:
#   invoice Date with Invoice_ID as key, AND
#   lists of Invoice_ID with invoice Date as key
dates_by_inv = {}
inv_by_dates = {}
for i in range(2,47+1):
        dates_by_inv[str(inv_wksheet.cell(i,2).value)] = inv_wksheet.cell(i,3).value
        if inv_wksheet.cell(i,3).value in inv_by_dates:
            inv_by_dates[str(inv_wksheet.cell(i,3).value)].append(int(inv_wksheet.cell(i,2).value))
        else:
            inv_by_dates[str(inv_wksheet.cell(i,3).value)] = [int(inv_wksheet.cell(i,2).value)]

print('\ninvoice Date with Invoice_ID as key')
print(dates_by_inv)
print('\nInvoice_ID with invoice Date as key')
print(inv_by_dates)


# Iterate over rows in worksheet, build dictionary with Invoice_ID as key.
# Each key has as value a list of dictionaries, with:
#   Product_ID as key & Units as value
line_items_by_inv = {}
for i in range(2,66+1):
        if line_items_wksheet.cell(i,1).value in line_items_by_inv:
                line_items_by_inv[line_items_wksheet.cell(i,1).value].append({str(line_items_wksheet.cell(i,2).value): int(line_items_wksheet.cell(i,3).value)})
        else:
                line_items_by_inv[line_items_wksheet.cell(i,1).value] = [{str(line_items_wksheet.cell(i,2).value): int(line_items_wksheet.cell(i,3).value)}]

print('\nlist of invoice line items with Invoice_ID as key')
print(line_items_by_inv)


# Iterate over rows in worksheet, build dictionary of:
#   unit Price with Product_ID as key
products_by_id = {}
for i in range(2,9+1):
        products_by_id[str(products_wksheet.cell(i,1).value)] = int(products_wksheet.cell(i,3).value)
        
print('\nunit Price with Product_ID as key')
print(products_by_id)
print('-----------------------------------\n\n')

for (k,v) in inv_by_dates.items():
    print(f'invoices dated {k}: {v}')
    for i in v:
        print(f'inv.# {i} line items: {line_items_by_inv[str(i)]}')
        for j in line_items_by_inv[str(i)]:
            print(j)
            for l in j.items():
                print(f'l: {l}')
                prod_id = str(l[0])
                units_sold = l[1]
                unit_price = products_by_id[prod_id]
                line_total = units_sold * unit_price

                print(f'prod. id: {prod_id}')
                print(f'# units sold: {units_sold}')
                print(f'unit price: {unit_price}')
                print(f'line total: {line_total}')
                
