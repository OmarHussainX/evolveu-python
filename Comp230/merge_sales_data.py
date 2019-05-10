from validate_sales_data import validate_sales_data
from pathlib import Path
import openpyxl
import pandas
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, NamedStyle

merged_file = 'merged_sales_data.xlsx'


def merge_sales_data(file1, file2):
    """
    Receives references to two sales data spreadsheets which contain
    validated data, and intelligently (checks for clients common to
    the two data sets and does not repeat them) merges them into a new
    spreadsheet
    """

    # ----------------------------------------------------------
    def copy_sheet(source, target, source_rows_max, source_cols_max,
                   source_start_row, target_start_offset):
        """
        Helper function
        Copies an entire worksheet from 'source' to 'target',
        which are in *different* workbooks (this precludes the use
        of 'target = wb.copy_worksheet(source)' which only works for
        sheets in the same workbook)
        'source_rows_max'       number of rows in 'source' worksheet
        'source_cols_max'       number of columns in 'source' worksheet
        'source_start_row'      1st row for data to be copied from 'source'
        'target_start_offset'   # of rows in 'target' to skip when copying data
        """
        for i in range(source_start_row, source_rows_max + 1):
            # NOTE: skipping blank rows found in the source sheet doesn't help:
            # Even if the blank source row is skipped, the row counter must
            # advance by one, which will by necessity create a blank row in the
            # target.
            for j in range(1, source_cols_max + 1):
                target.cell(target_start_offset + i, j).value = \
                    source.cell(i, j).value
    # ----------------------------------------------------------

    # Create new (target) workbook with blank sheets
    wb_new = openpyxl.Workbook()
    ws_clients = wb_new.active
    ws_clients.title = 'customers'
    ws_invoices = wb_new.create_sheet('invoices')
    ws_line_items = wb_new.create_sheet('invoice line items')
    ws_products = wb_new.create_sheet('products')

    # Open first workbook whose data is to be merged, and then...
    wb = openpyxl.load_workbook(file1, data_only=True)

    # ...copy all worksheets of interest from first workbook to target
    copy_sheet(wb['customers'],
               ws_clients,
               wb['customers'].max_row,
               wb['customers'].max_column,
               1, 0)

    copy_sheet(wb['invoices'],
               ws_invoices,
               wb['invoices'].max_row,
               wb['invoices'].max_column,
               1, 0)

    copy_sheet(wb['invoice line items'],
               ws_line_items,
               wb['invoice line items'].max_row,
               wb['invoice line items'].max_column,
               1, 0)

    copy_sheet(wb['products'],
               ws_products,
               wb['products'].max_row,
               wb['products'].max_column,
               1, 0)

    # Build list of customer IDs from first workbook, and:
    # - remove header cell ('Customers')
    # - remove values from empty cells (None)
    file1_customers = [cell.value for cell in wb['customers']['A']]
    del file1_customers[0]
    file1_customers = [id for id in file1_customers if id is not None]

    # Open second workbook whose data is to be merged, and then...
    wb = openpyxl.load_workbook(file2, data_only=True)

    # ...copy all worksheets of interest from second workbook to target.
    # Other worksheets can be merged into the target worksheet using the
    # 'copy_sheet' helper function, but the 'customers' worksheet needs to
    # be handled separately, taking care to:
    #   - skip the first row of the source (headers already present in target)
    #   - start writing data from the *last* (not first!) row of the target
    #   - ensure that customers which exist in the first workbook *and* the
    #     second workbook are not copied to the merged workbook twice
    merge_row_offset = ws_clients.max_row
    for i in range(2, wb['customers'].max_row + 1):
        if wb['customers'].cell(i, 1).value in file1_customers:
            continue
        for j in range(1, wb['customers'].max_column + 1):
                ws_clients.cell(merge_row_offset + i, j).value = \
                    wb['customers'].cell(i, j).value

    copy_sheet(wb['invoices'],
               ws_invoices,
               wb['invoices'].max_row,
               wb['invoices'].max_column,
               2, ws_invoices.max_row)

    copy_sheet(wb['invoice line items'],
               ws_line_items,
               wb['invoice line items'].max_row,
               wb['invoice line items'].max_column,
               2, ws_line_items.max_row)

    # ----------------------------------------------------------
    def worksheet_cleaner(worksheet, key_header_title, sheet_index):
        """
        Helper function
        Receives a reference to a worksheet in merged workbook 'wb_new', and
        the header title of a key column in that worksheet. Deletes the
        worksheet and creates a new one in 'wb_new' at 'sheet_index' which is:
        - sorted by the key header
        - devoid of any rows where the key column is empty/NaN
        """
        data = worksheet.values
        # 'data' is an iterable generator object -
        # get column headers from first row
        columns = next(data)

        # Create a DataFrame from the second row of data onward, and:
        # - sort by key header (this also collects blank rows at the end)
        # - delete rows iwhere the key column is NaN/empty
        df = pandas.DataFrame(data, columns=columns)
        df.sort_values([key_header_title], inplace=True)
        df.dropna(subset=[key_header_title], inplace=True)

        # remove current worksheet from merged workbook, and...
        wb_new.remove(worksheet)

        # ...add new worksheet with the same title, and...
        worksheet = wb_new.create_sheet(worksheet.title, sheet_index)

        # ...append the data from the sorted and cleaned DataFrame
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

    # ----------------------------------------------------------

    # 'Clean' each worksheet in the merged workbook:
    # - sort by ID, and remove blank rows
    worksheet_cleaner(ws_clients, 'Customer', 0)
    worksheet_cleaner(ws_invoices, 'Invoice', 1)
    worksheet_cleaner(ws_line_items, 'Line', 2)

    # Create styles to use to format the cells of the merged workbook
    default_cell = NamedStyle(name='default_cell')
    default_cell.font = Font(name='Arial', size=12)
    wb_new.add_named_style(default_cell)

    bold_header = NamedStyle(name='bold_header')
    bold_header.font = Font(name='Arial', size=13, bold=True)
    wb_new.add_named_style(bold_header)

    date_cell = NamedStyle(name='date_cell', number_format='YYYY-MM-DD')
    date_cell.font = Font(name='Arial', size=12)
    wb_new.add_named_style(date_cell)

    # Apply cell styles, adjust row height aand column width
    for sheet in wb_new:
        sheet.sheet_format.defaultColWidth = 14
        sheet.sheet_format.defaultRowHeight = 18
        for row in sheet.iter_rows():
            for cell in row:
                cell.style = 'default_cell'
                if cell.row == 1:       # make header bold
                    cell.style = 'bold_header'

    # Grab second column from 'invoices' worksheet and format as date
    invoices_date_col = wb_new['invoices']['B']
    for cell in invoices_date_col:
        if cell.row > 1:                # skip header
            cell.style = 'date_cell'

    # Save merged workbook to disk
    wb_new.save(merged_file)


def main():
    """
    Prompt user to enter filenames for two sales data spreadsheets,
    which will be merged into a new spreadsheet
    """

    print(f'\n== This will merge two sales data spreadsheets into \
\'{merged_file}\' ==\n')

    file1 = input("Enter first file (hit enter for 'sales_data.xlsx')\n> ")
    if file1 == '':
        file1 = 'sales_data.xlsx'
    print(f'First file: \'{file1}\'\n')

    file_path1 = Path(file1)
    if not file_path1.is_file():
        print(f'ERROR: file \'{file1}\' not found')
        return

    if not validate_sales_data(file1):
        print(f'ERROR: file \'{file1}\' contains invalid data\n\
See \'validate_sales_data_report.txt\' for details')
        return

    file2 = input("Enter second file (hit enter for 'sales_data2.xlsx')\n> ")
    if file2 == '':
        file2 = 'sales_data2.xlsx'
    print(f'Second file: \'{file2}\'\n')

    file_path2 = Path(file2)
    if not file_path2.is_file():
        print(f'ERROR: file \'{file2}\' not found')
        return

    if not validate_sales_data(file2):
        print(f'ERROR: file \'{file2}\' contains invalid data\n\
See \'validate_sales_data_report.txt\' for details')
        return

    merge_sales_data(file1, file2)
    print(f'Merged \'{file1}\' and \'{file2}\' into \'{merged_file}\'')

if __name__ == '__main__':
    main()
