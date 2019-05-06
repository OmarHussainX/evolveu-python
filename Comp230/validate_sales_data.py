import openpyxl


def validate_sales_data(filename):
    """
    sales_data.xlsx MUST have four worksheets (in any order) named:
    'customers', 'invoices', 'invoice line items', 'products'

    Spreadsheet data will be validated to ensure that it has:
    * 10 - 15 _unique_ clients
    * 3 - 4 invoices per client
    * 1 - 5 items per invoice
    * invoices for one month only
    * $15,000 (+/- 1,500) of invoices per month
    """

    # Set a bunch of flags, check them at the end to determine whether
    # or not the data set is valid
    clients_not_repeated = True
    client_count_in_range = True
    client_invoice_count_in_range = True
    line_items_in_range = True
    invoices_for_one_month = True
    invoices_total_in_range = True

    # Load worksheet iterators
    # Set 'data_only' so that the _results_ of formulae are accessible
    wb = openpyxl.load_workbook(filename, data_only=True)
    client_sheet = wb['customers']
    invoices_sheet = wb['invoices']
    line_items_sheet = wb['invoice line items']

    # Create list of client #
    # - check for duplicates
    # - check if there are between 10 and 15 clients
    client_list = []
    for col in client_sheet.iter_cols(min_row=2, max_col=1,
                                      max_row=client_sheet.max_row):
        for cell in col:
            if not (cell.value is None):
                client_list.append(cell.value)

    clients_not_repeated = len(client_list) == len(set(client_list))
    client_count_in_range = 10 <= len(client_list) <= 15

    # Ensure there are 3 - 4 invoices per client:
    # - using client # as key, build up a list of invoices, and then
    #   check that each client has 3 - 4 invoices
    # Ensure all invoices were issued in the same month:
    # - build up a list of the month of issue for each invoice, and then
    #   check that all invoices were issued in the same month
    invoices_per_client = {}
    items_per_invoice = {}
    invoices_month = []
    for i in range(2, invoices_sheet.max_row + 1):
            if not (invoices_sheet.cell(i, 3).value is None):
                if invoices_sheet.cell(i, 3).value in invoices_per_client:
                        invoices_per_client[invoices_sheet.cell(i, 3).value]\
                            .append(invoices_sheet.cell(i, 1).value)
                else:
                        invoices_per_client[invoices_sheet.cell(i, 3).value] =\
                            [invoices_sheet.cell(i, 1).value]

                # create key for invoice #
                items_per_invoice[invoices_sheet.cell(i, 1).value] = []

                invoices_month.append(invoices_sheet.cell(i, 2).value
                                      .strftime("%B"))

    invoices_for_one_month = len(set(invoices_month)) == 1

    for invoices in invoices_per_client.values():
        if len(invoices) < 3 or len(invoices) > 4:
            client_invoice_count_in_range = False

    # Ensure there are 1 - 5 items per invoice:
    # - using invoice # as key, build up a list of line items, and then
    #   check that each invoice has 1 - 5 line items
    # Ensure total of all invoices is $15,000 (+/- 1,500):
    # - add up the line totals for each invoice to get the total of all
    #   invoices issued, and then ensure it is valid
    invoices_total = 0
    for i in range(2, line_items_sheet.max_row + 1):
            if not (line_items_sheet.cell(i, 2).value is None):
                # keys for each invoice # were created earlier
                if line_items_sheet.cell(i, 2).value in items_per_invoice:
                    items_per_invoice[line_items_sheet.cell(i, 2).value]\
                                .append(line_items_sheet.cell(i, 3).value)

                invoices_total += line_items_sheet.cell(i, 6).value

    invoices_total_in_range = (15000-1500) <= invoices_total <= (15000+1500)

    for line_items in items_per_invoice.values():
        if len(line_items) < 1 or len(line_items) > 5:
            line_items_in_range = False

    # Generate report
    data_validation_result = True
    with open('validate_sales_data_report.txt', 'w') as report:
        if clients_not_repeated and client_count_in_range and \
            client_invoice_count_in_range and invoices_for_one_month and \
                invoices_total_in_range and line_items_in_range:
            report.write(f'{filename} - data VALIDATED\n\n')
        else:
            data_validation_result = False
            report.write(f'{filename} - data INVALID\n\n')

        report.write(format('10 - 15 clients', '40') +
                     str(client_count_in_range) + '\n')
        report.write(format('clients are unique', '40') +
                     str(clients_not_repeated) + '\n')
        report.write(format('3 - 4 invoices per client', '40') +
                     str(client_invoice_count_in_range) + '\n')
        report.write(format('1 - 5 items per invoice', '40') +
                     str(line_items_in_range) + '\n')
        report.write(format('invoices in same month', '40') +
                     str(invoices_for_one_month) + '\n')
        report.write(format('invoices total = $15,000 (+/- 1,500)', '40') +
                     str(invoices_total_in_range) + '\n')

    return data_validation_result


def main():
    # print(f'\n---------- {__file__} ----------')
    validate_sales_data('sales_data.xlsx')


if __name__ == '__main__':
    main()
